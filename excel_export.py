"""Build a formatted Excel workbook from exported CSV."""

from __future__ import annotations

import csv
from pathlib import Path

DEFAULT_CSV_NAME = "extracted_ris_data.csv"
DEFAULT_XLSX_NAME = "extracted_ris_data.xlsx"

# Keep wide columns readable in Excel without stretching the sheet too far.
MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 48
CELL_PADDING = 2


def csv_to_formatted_xlsx(
    csv_path: Path,
    xlsx_path: Path | None = None,
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    csv_path = Path(csv_path)
    if not csv_path.is_file():
        raise FileNotFoundError(f"CSV not found: {csv_path}")

    target = Path(xlsx_path) if xlsx_path else csv_path.with_name(DEFAULT_XLSX_NAME)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "References"

    with csv_path.open(newline="", encoding="utf-8-sig") as handle:
        for row_index, row in enumerate(csv.reader(handle), start=1):
            for col_index, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.alignment = Alignment(wrap_text=row_index > 1, vertical="top")

    header_fill = PatternFill("solid", fgColor="E6F4F1")
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        header_len = len(str(column_cells[0].value or ""))
        max_content = header_len
        for cell in column_cells[1:]:
            if cell.value is None:
                continue
            text = str(cell.value).replace("\r\n", "\n")
            longest_line = max((len(part) for part in text.split("\n")), default=0)
            max_content = max(max_content, longest_line)
        sheet.column_dimensions[letter].width = max(
            MIN_COLUMN_WIDTH,
            min(max_content + CELL_PADDING, MAX_COLUMN_WIDTH),
        )

    workbook.save(target)
    return target
